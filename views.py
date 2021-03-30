from django.shortcuts import render
from .models import *
from .forms import loginForm,PracownikFormConf,LoginForm2
from django.http import HttpResponseRedirect,HttpResponse
from .functions  import aes_it,deaes_it
from .forms import SignatureForm
from jsignature.utils import draw_signature
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
#import pymysql.cursors
#from django.http import httpresponse

def start(request):
    return None
#NARZEDZIA
@login_required(login_url='/narzedziownia/login2/')
def narzedzia_lista(request, *args, **kwargs):
    context = {
    }
    url='narzedziownia/narzedzia_lista.html'
    Narzedzia=Narzedzie.objects.all()
    print(Narzedzia)
    context['Narzedzia']=Narzedzia   
    return render(request,url,context)
@login_required(login_url='/narzedziownia/login2/')
def narzedzia_edit(request, *args, **kwargs):
    print(request.POST)
    context = {
    }
    url='narzedziownia/narzedzia_edit.html'
    pk=kwargs['pk']
    Narz=Narzedzie.objects.get(pk=pk)
    if request.method == 'POST':
        form = NarzedzieForm(request.POST,instance=Narz)
        if form.is_valid():
            Narz.nazwa=form.cleaned_data['nazwa']
            Narz.opis=form.cleaned_data['opis']
            Narz.save()
            return HttpResponseRedirect("/narzedziownia/narzedzia_lista/")
        else:
            None
    else:
        form = NarzedzieForm()
    context['form']=form
    context['Narz']=Narz
       #context['form']=form
    return render(request,url,context)
def narzedzia_usun(request, *args, **kwargs):
    pk=kwargs['pk']

    if Pobranie.objects.filter(narzedzie=pk).count() == 0:

        Narzedzie.objects.get(pk=pk).delete()
    else:
        return HttpResponseRedirect("/narzedziownia/narzedzia_usun_alert/"+pk+"")

    return HttpResponseRedirect("/narzedziownia/narzedzia_lista/")
def narzedzia_usun_alert(request, *args, **kwargs):
    url='narzedziownia/narzedzia_usun_alert.html'
    pk = kwargs['pk']
    Narz = Narzedzie.objects.get(pk=pk)
    Pobrania = Pobranie.objects.filter(narzedzie=pk)
    context = { 'Pobrania':Pobrania, 'Narz':Narz }

    return render(request,url,context)
    
def narzedzia_new(request, *args, **kwargs):
    print(request.POST)
    context = {
    }
    url='narzedziownia/narzedzia_new.html'
    Narz=Narzedzie()
    if request.method == 'POST':
        form = NarzedzieForm(request.POST)
        if form.is_valid():
            print("ffff")
            print(form.cleaned_data['nazwa'])

            Narz.nazwa=form.cleaned_data['nazwa']
            Narz.opis=form.cleaned_data['opis']
            Narz.save()
            return HttpResponseRedirect("/narzedziownia/narzedzia_lista/")
        else:
            None
    else:
        form = NarzedzieForm()
    context['form']=form
    context['Narz']=Narz
    context['var']=1
       #context['form']=form
    return render(request,url,context)

#ODZIEZ
@login_required(login_url='/narzedziownia/login2/')
def odziez_lista(request, *args, **kwargs):
    context = {
    }
    if request.method == 'GET' and request.GET!={}:
        pozostaja=request.GET
        SzablonOdziez.objects.all().exclude(pk__in=pozostaja).delete()
        for i in pozostaja:
            print(i)
            o=Odziez.objects.get(pk=i)
            SzablonOdziez.objects.get_or_create(dzial=".",wariant=".",odziez=o,ilosc=1)


        

    url='narzedziownia/odziez_lista.html'
    SzOdz=SzablonOdziez.objects.all()
    Odz=Odziez.objects.all()
    context['Odziez']=Odz
    context['SzablonOdziez']=SzOdz
    return render(request,url,context)
@login_required(login_url='/narzedziownia/login2/')
def odziez_edit(request, *args, **kwargs):
    print(request.POST)
    context = {
    }
    url='narzedziownia/odziez_edit.html'
    pk=kwargs['pk']
    Odz=Odziez.objects.get(pk=pk)
    if request.method == 'POST':
        form = OdziezForm(request.POST,instance=Odz)
        if form.is_valid():
            Odz.nazwa=form.cleaned_data['nazwa']
            Odz.opis=form.cleaned_data['opis']
            Odz.save()
            return HttpResponseRedirect("/narzedziownia/odziez_lista/")
        else:
            None
    else:
        form = OdziezForm()
    context['form']=form
    context['Odziez']=Odz
       #context['form']=form
    return render(request,url,context)
def odziez_usun(request, *args, **kwargs):
    pk=kwargs['pk']
    Odziez.objects.get(pk=pk).delete()
    return HttpResponseRedirect("/narzedziownia/odziez_lista/")
def odziez_new(request, *args, **kwargs):
    print(request.POST)
    context = {
    }
    url='narzedziownia/odziez_new.html'
    Odz=Odziez()
    if request.method == 'POST':
        form = OdziezForm(request.POST)
        if form.is_valid():
            print("ffff")
            print(form.cleaned_data['nazwa'])

            Odz.nazwa=form.cleaned_data['nazwa']
            Odz.opis=form.cleaned_data['opis']
            Odz.save()
            return HttpResponseRedirect("/narzedziownia/odziez_lista/")
        else:
            None
    else:
        form = NarzedzieForm()
    context['form']=form
    context['Odziez']=Odz
    context['var']=1
       #context['form']=form
    return render(request,url,context)
#EMPLOYEE/PRACOWNICY

#EMPLOYEE/PRACOWNICY
#@login_required

@login_required(login_url='/narzedziownia/login2/') 
def pracownicy_lista(request, *args, **kwargs):
    import MySQLdb
    #=== DODAJ ZATRUDNIENIA
    try:
        conn=MySQLdb.connect(host="192.168.41.15",user="kadry", passwd="start", db="KADRY",port=3306,charset='utf8')
        cur = conn.cursor()
        sql = "SELECT * FROM Nowi_pracownicy" 
        cur.execute(sql)
        rows=cur.fetchall()
    except mdb.Error:
        print("ERROR")
        #print "Error %d: %s" % (e.args[0],e.args[1])
        sys.exit(1)
    finally:
        if conn:
            conn.close()
    for i in rows:
        Prac, created=Pracownik.objects.get_or_create(nazwisko_imie="{} {}".format(i[2],i[1]),dzial="{}".format(i[3]).upper(),stanowisko="{}".format(i[4]))

        #Prac=Pracownik()
        #Prac.nazwisko_imie="{} {}".format(i[2],i[1])
        #Prac.dzial="{}".format(i[3])
        #Prac.stanowisko="{}".format(i[4])
        #Prac.save()
        
        #print(i[0])
        #print(i[1])
        #print(i[2])
        #print(i[3])
        #print(i[4])
        
        

    context = {
    }
    url='narzedziownia/pracownicy_lista.html'
    Pracownicy=Pracownik.objects.filter(zwolniony=False)
    context['Pracownicy']=Pracownicy
    return render(request,url,context)

def pracownicy_lista_zwolnionych(request, *args, **kwargs):

        #Prac=Pracownik()
        #Prac.nazwisko_imie="{} {}".format(i[2],i[1])
        #Prac.dzial="{}".format(i[3])
        #Prac.stanowisko="{}".format(i[4])
        #Prac.save()
        
        #print(i[0])
        #print(i[1])
        #print(i[2])
        #print(i[3])
        #print(i[4])
        
    context = {
    }
    url='narzedziownia/pracownicy_lista_zwolnionych.html'
    Pracownicy=Pracownik.objects.filter(zwolniony=True)
    context['Pracownicy']=Pracownicy
    return render(request,url,context)
def pracownicy_edit_(request, *args, **kwargs):
    from collections import OrderedDict
    context = {
    }
    url='narzedziownia/pracownicy_edit.html'
    pk=kwargs['pk']
    #print(kwargs)
    Numery_narz=range(10)
    Prac=Pracownik.objects.get(pk=pk)
    Narz_temp_str=Prac.narzedzia
    if Narz_temp_str!=None:
        Narz_temp_list=Narz_temp_str.split(";")
        Narz_temp=OrderedDict() 
        print(Narz_temp_list)

        for i in Narz_temp_list:
            if i.split(":")[0]!="":  #wuwalamy ostatniego pustaka
                try:
                    Narz_temp[i.split(":")[0]]=i.split(":")[1]
                except:
                    break;
    else:
        Narz_temp=[]
    

    Lista_narz=Narzedzie.objects.all()
    print(request.POST)
    narz_list=""
    if request.method == 'POST':
        form = PracownikForm(request.POST,instance=Prac)
        if form.is_valid():
            Prac.komentarz=form.cleaned_data['komentarz']
            for i in Numery_narz:
                narz="narz_{}".format(i)
                #is_narz="is_narz_{}".format(i)
                il_narz="il_narz_{}".format(i)
                narz_list+=request.POST[narz]
                narz_list+=":"
                narz_list+=request.POST[il_narz]
                narz_list+=";"
            Prac.narzedzia=narz_list
            Prac.save()
            return HttpResponseRedirect("/narzedziownia/pracownicy_lista/")
        else:
            None
    else:
        form =PracownikForm()
    context['form']=form
    context['Prac']=Prac
    context['Numery_narz']=Numery_narz
    


    context['Lista_narz']=Lista_narz
    context['Narz_temp']=Narz_temp
       #context['form']=forma
    print(Lista_narz,Narz_temp)
    return render(request,url,context)

def pracownicy_usun_pobranie(request, *args, **kwargs):
    #context = {
    #}
    #url='narzedziownia/pracownicy_usun_pobranie.html'
    pk_pobr=kwargs['pk_pobr']
    pk_prac=kwargs['pk_prac']
    Pobr=Pobranie.objects.filter(pk=pk_pobr).delete()
    return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk_prac+"/")
def pracownicy_usun_pobranie_barcode(request, *args, **kwargs):
    #context = {
    #}
    #url='narzedziownia/pracownicy_usun_pobranie.html'
    pk_pobr=kwargs['pk_pobr']
    pk_prac=kwargs['pk_prac']
    Pobr=Pobranie.objects.filter(pk=pk_pobr).delete()
    return HttpResponseRedirect("/narzedziownia/pracownicy_edit_barcode/"+pk_prac+"/")

@login_required(login_url='/narzedziownia/login2/')
def pracownicy_edit(request, *args, **kwargs):
    from datetime import datetime
    from django.urls import resolve
    context = {
    }
    dzialy = Pracownik.objects.values('dzial').distinct()
    warianty = Szablon.objects.values('wariant','dzial').distinct()
    
    url='narzedziownia/pracownicy_edit.html'
    pk=kwargs['pk']
    #print(kwargs)
    current_url = resolve(request.path_info).url_name
    print(current_url)
    Prac=Pracownik.objects.get(pk=pk)
    if Prac.zwolniony == True and current_url != "pracownicy_edit_arch":
        return HttpResponseRedirect("/narzedziownia/zwolniony_alert/"+pk+"/")
    Narz_temp_str=Prac.narzedzia

    Lista_narz=Narzedzie.objects.all()
    narz_list=""
    Pobr=Pobranie()
    # Add sth invisible, like pk in oder to avoid sending empty GET after submitting.
    pozostaja=[]
    if request.method == 'GET' and request.GET!={}:
        for i in request.GET:
            if i.startswith("opis"):
                p = Pobranie.objects.get(pk=i[4:])
                p.opis=request.GET[i]
                p.save()
            elif i.startswith("data"):
                p = Pobranie.objects.get(pk=i[4:])
                p.data_pobrania=request.GET[i]
                p.save()

            else:
                pozostaja.append(i)

        Pobranie.objects.filter(pracownik=Prac).exclude(pk__in=pozostaja).delete()
        return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk+"/")
    if request.method == 'POST':
        #print(request.POST['opis'])
        try:
            pk_narz=request.POST['narzedzie']
            Pobr.pracownik=Prac
            Pobr.narzedzie=Narzedzie.objects.get(pk=pk_narz)
            Pobr.opis=request.POST['opis']
            Pobr.ilosc=request.POST['ilosc']
            if request.POST['data_pobrania']!="":
                Pobr.data_pobrania=request.POST['data_pobrania']

            ### USUNIETA DATA ODDANIA
            #if request.POST['data_oddania']!="" and  request.POST['data_oddania']!=None:
            #    Pobr.data_oddania=request.POST['data_oddania']
            Pobr.save()
            return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk+"/")
        except:
            ## USUWAMY TYLKO TYCH KTORYCH NIE MA
            try:
                l=list(request.POST.keys())
                l=list(filter(lambda x:not x.startswith(("d","n","o")) ,l))
                l2=[]
                for i in l:
                    try:
                        k=int(i)
                        l2.append(k)
                    except Exception as a:
                        print(a)
                PobranieOdziez.objects.filter(pracownik=Prac).exclude(odziez__pk__in=l2).delete()
                #Nie dziala exclude
            except Exception as a:
                print(a)
            ## DODAJEMY TYCH KTORYCH NIE MA....

            l=list(request.POST.keys())
            lc=list(filter(lambda x:not x.startswith(("d","n","o")) ,l))
            ld=list(filter(lambda x:x.startswith("d") ,l)) ####DATY 
            ln=list(filter(lambda x:x.startswith("n") ,l)) ####Ilosci
            lo=list(filter(lambda x:x.startswith("o") ,l)) ####Oddano trzeba dorobic HTML
            for k in lc:
                try:
                    
                    Od=Odziez.objects.get(pk=k)

                    P=PobranieOdziez.objects.get_or_create(odziez=Od,pracownik=Prac)
                    
                    #P.data_pobrania=
                    
                except Exception as a:
                    print(a)
            #UZUPENILAJ DATY
            for k in ld:
                p=k[1:]
                
                O=Odziez.objects.get(pk=p)
                #print(O)
                #print(Od)
                try:
                    P=PobranieOdziez.objects.get(odziez=O,pracownik=Prac)
                    P.data_pobrania=request.POST[k]
                    P.save()
                except:
                    None
            #UZUPELNIAJ ILOSCI
            for k in ln:
                p=k[1:]
            #    
                O=Odziez.objects.get(pk=p)
                print(O)
                #print(Od)
                try:
                    P=PobranieOdziez.objects.get(odziez=O,pracownik=Prac)
                    P.ilosc=request.POST[k]
                    P.save()
                except:
                    None
            #Oddano 29032021
            for k in lo:
                p=k[1:]
            #    
                O=Odziez.objects.get(pk=p)
                print(O)
                #print(Od)
                try:
                    P=PobranieOdziez.objects.get(odziez=O,pracownik=Prac)
                    P.oddano=request.POST[k]
                    P.save()
                except:
                    None


            #    data_pobrania= request.POST[k]
           #:     print("aaaaaaaa",pk)
                #P=PobranieOdziez.objects.get(pk=int(pk))
                #P.data_pobrania=data_pobrania
                #P.save()




                #print(Odz)

            #print(odziez)
            #for odz in odziez:
            #    print("MODAL")
            #    print(odz)
            form = PobranieForm() #Może jakoś mądzej obejsc ten problem????
        #else:
        #    None
    else:
        form = PobranieForm()
        #form =PracownikForm()a
    Pobr_filter=Pobranie.objects.filter(pracownik=pk)
    Odz = list(SzablonOdziez.objects.order_by('odziez__nazwa').values())
    #Odz = list(SzablonOdziez.objects.values()) # Powyzej zmiania z 22-03-2021 sort ubran w modalu
    for o in Odz:
        n=Odziez.objects.get(pk=o['odziez_id']).nazwa

        #n=o.Odziez.nazwa
        o['nazwa']=n
        pk_odziez=o['odziez_id']
        if PobranieOdziez.objects.filter(pracownik=pk,odziez__pk=pk_odziez):
            o['checked']=True
            o['data_pobrania']=PobranieOdziez.objects.get(pracownik=pk,odziez__pk=pk_odziez).data_pobrania
            o['ilosc']=PobranieOdziez.objects.get(pracownik=pk,odziez__pk=pk_odziez).ilosc
        else:
            o['checked']=False
            #o['data_pobrania']=PobranieOdziez.objects.get(pracownik=pk,odziez__pk=pk_odziez).data_pobrania
    #print(Odz)
    PobranieO = PobranieOdziez.objects.filter(pracownik=Prac)
    context['PobranieOdziez']=PobranieO
    context['Odziez']=Odz
    context['Pobr_filter']=Pobr_filter
    context['form']=form
    context['Prac']=Prac
    context['Lista_narz']=Lista_narz
    context['dzialy']=dzialy
    context['warianty']=warianty
       #context['form']=forma
    
    return render(request,url,context)

def pracownicy_view(request, *args, **kwargs):
    from datetime import datetime
    context = {
    }
    url='narzedziownia/pracownicy_view.html'
    pk=kwargs['pk']
    #print(kwargs)
    Prac=Pracownik.objects.get(pk=pk)
    Narz_temp_str=Prac.narzedzia
    b=b''

    Lista_narz=Narzedzie.objects.all()
    narz_list=""
    Pobr=Pobranie()
    Pobr_filter=Pobranie.objects.filter(pracownik=pk)
    PobrOdziez_filter=PobranieOdziez.objects.filter(pracownik=pk)
    if request.method == 'GET':
        form_s = SignatureForm(request.GET)
        form = PracownikFormConf(request.GET)
        if form.is_valid():
            ss = form.cleaned_data['potwierdzenie']
            if str(ss) == str(Prac.nr_karty): #  form.cleaned_data['potwierdzenie'] == str(Prac.nr_karty):
                print(aes_it(form.cleaned_data['potwierdzenie']))
             
                sgn=form.cleaned_data['potwierdzenie'] +datetime.now().strftime("%d:%m:%Y:%H:%M:%S")
                print(sgn)
                for p in Pobr_filter:
                    p.signature=aes_it(sgn)
                    p.save()
                for po in PobrOdziez_filter:
                    po.signature=aes_it(sgn)
                    po.save()
                #Prac.save()
                #for p in Pobr_filter:
                #    print(p.signature)
                return HttpResponseRedirect("/narzedziownia/confirmed/")
            else:
                return HttpResponseRedirect("/narzedziownia/notconfirmed/")
        elif form_s.is_valid():
            for p in Pobr_filter:
                if p.signature_handy == None:
                    p.signature_handy=form_s.cleaned_data['signature']
                    p.save()
            ##Poprawka 02-03-2021
            for po in PobrOdziez_filter:
                if po.signature_handy == None:
                    po.signature_handy=form_s.cleaned_data['signature']
                    po.save()
            #Prac.save()
            for p in Pobr_filter:
                print(p.signature_handy)
            return HttpResponseRedirect("/narzedziownia/confirmed/")

        else:
            None
    else:
        form =PracownikFormConf()
    for i in Pobr_filter:
        try:
            b=memoryview(i.signature).tobytes()
            print(deaes_it(b))
        except:
            None
    context['form_s']=form_s
    context['form']=form
    context['Prac']=Prac
    Pobr_filter_kol1=Pobranie.objects.filter(pracownik=pk)[:9]
    Pobr_filter_kol2=Pobranie.objects.filter(pracownik=pk)[9:19]
    Pobr_filter_kol3=Pobranie.objects.filter(pracownik=pk)[19:]
    context['PobrOdziez_filter']=PobrOdziez_filter
    context['Pobr_filter']=Pobr_filter
    context['Pobr_filter_kol1']=Pobr_filter_kol1
    context['Pobr_filter_kol2']=Pobr_filter_kol2
    context['Pobr_filter_kol3']=Pobr_filter_kol3
    return render(request,url,context)
@login_required(login_url='/narzedziownia/login2/')
def pracownicy_edit_barcode(request, *args, **kwargs):
    from datetime import date
    context = {
    }
    url='narzedziownia/pracownicy_edit_barcode.html'
    pk=kwargs['pk']
    Prac=Pracownik.objects.get(pk=pk)
    komunikat=""
    if request.method == 'POST':
        pk_narz=request.POST['barcode']
        try:
            P=Pobranie()
            Narz=Narzedzie.objects.get(pk=pk_narz)
            print(Narz)
            P.pracownik=Prac
            P.narzedzie=Narz
            P.data_pobrania=date.today()
            P.ilosc=1
            P.save()
            komunikat="dodano"
        except:
            komunikat="brak narzedzia"
            

    Pobr_filter=Pobranie.objects.filter(pracownik=pk)
    context['Pobr_filter']=Pobr_filter
    context['Prac']=Prac
    context['komunikat']=komunikat
    return render(request,url,context)

def pracownicy_view_(request, *args, **kwargs):
    from collections import OrderedDict
    context = {
    }
    url='narzedziownia/pracownicy_view.html'
    pk=kwargs['pk']
    #print(kwargs)
    Numery_narz=range(10)
    Prac=Pracownik.objects.get(pk=pk)
    Narz_temp_str=Prac.narzedzia
    if Narz_temp_str!=None:
        Narz_temp_list=Narz_temp_str.split(";")
        Narz_temp=OrderedDict() 
        print(Narz_temp_list)

        for i in Narz_temp_list:
            if i.split(":")[0]!="":  #wuwalamy ostatniego pustaka
                try:
                    Narz_temp[i.split(":")[0]]=i.split(":")[1]
                except:
                    break;
    else:
        Narz_temp=[]
    

    Lista_narz=Narzedzie.objects.all()
    print(request.POST)
    narz_list=""
    if request.method == 'GET':
        form = PracownikFormConf(request.GET)
        if form.is_valid():
            print(form.cleaned_data['potwierdzenie'])
            Prac.potwierdzenie=form.cleaned_data['potwierdzenie']

            #Prac.potwierdzenie=request.POST('potwierdzenie')
            
            #Prac.komentarz=form.cleaned_data['komentarz']
            #for i in Numery_narz:
            #    narz="narz_{}".format(i)
            #    #is_narz="is_narz_{}".format(i)
            #    il_narz="il_narz_{}".format(i)
            #    narz_list+=request.POST[narz]
            #    narz_list+=":"
            #    narz_list+=request.POST[il_narz]
            #    narz_list+=";"
            #Prac.narzedzia=narz_list
            

            Prac.save()
            return HttpResponseRedirect("/narzedziownia/login/")
        else:
            None
    else:
        form =PracownikFormConf()
    context['form']=form
    context['Prac']=Prac
    context['Numery_narz']=Numery_narz
    


    context['Lista_narz']=Lista_narz
    context['Narz_temp']=Narz_temp
       #context['form']=form
    return render(request,url,context)

def pracownicy_arch(request, *args, **kwargs):
    prac=kwargs['pk']
    p=Pracownik.objects.get(pk=prac)
    p.zwolniony=True
    p.save()

    
    

    return HttpResponseRedirect("/narzedziownia/pracownicy_lista/")
def import_pracownikow(request, *args, **kwargs):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    import os
    url='narzedziownia/import_pracownikow.html'
    context={
    }

    path_src="Stan.XLSX"
    wb_src = load_workbook(path_src) 
    sheet_src = wb_src.active
    rows=[]
    
    for i in range(1500)[2:]:
        row = []
        cell_src_1 = sheet_src.cell(row = i, column = 1)
        cell_src_2 = sheet_src.cell(row = i, column = 2)
        cell_src_3 = sheet_src.cell(row = i, column = 3)
        cell_src_4 = sheet_src.cell(row = i, column = 4)
        cell_src_5 = sheet_src.cell(row = i, column = 5) #DATA
        cell_src_6 = sheet_src.cell(row = i, column = 6) #SPOT
        cell_src_7 = sheet_src.cell(row = i, column = 7) #Rodzaj wynagrodzenia
        cell_src_8 = sheet_src.cell(row = i, column = 8) #TypUmowy
        cell_src_9 = sheet_src.cell(row = i, column = 9) #Karta
        try:
            p=Pracownik.objects.get(nr_pracownika=cell_src_1.value)
            p.nr_pracownika=cell_src_1.value
            p.nazwisko_imie=cell_src_2.value
            p.dzial=cell_src_3.value
            p.stanowisko=cell_src_4.value
            p.zatrudnienie=cell_src_5.value
            p.spot=cell_src_6.value
            p.podgrupa_prac=cell_src_7.value
            p.typ_umowy=cell_src_8.value
            p.nr_karty=int(cell_src_9.value)
            p.save()
            rows.append(p)
        except:
            try:
                p,created=Pracownik.objects.get_or_create(nazwisko_imie=cell_src_2.value)
                p.nr_pracownika=cell_src_1.value
                p.nazwisko_imie=cell_src_2.value
                p.dzial=cell_src_3.value
                p.stanowisko=cell_src_4.value
                p.zatrudnienie=cell_src_5.value
                p.spot=cell_src_6.value
                p.podgrupa_prac=cell_src_7.value
                p.typ_umowy=cell_src_8.value
                p.nr_karty=int(cell_src_9.value)
                p.save()
                rows.append(p)
            except:
                print(cell_src_2.value)
        finally:
            
            None

        #    None
    context['rows']=rows
    return render(request,url,context)
def pracownicy_szablon(request, *args, **kwargs):
    from django.db.models import Q
    pk_prac = kwargs['pk_prac']
    Prac=Pracownik.objects.get(pk=pk_prac)
    dzial = kwargs['dzial']
    wariant = kwargs['wariant']
    wariant_=wariant+" "
    #Dzial=Prac.dzial
    #for s in Szablon.objects.filter(dzial=dzial,wariant__iregex=wariant_r):
    for s in Szablon.objects.filter(Q(dzial=dzial,wariant=wariant)|Q(dzial=dzial,wariant=wariant_)):
        Pobr =  Pobranie()
        Pobr.ilosc=s.ilosc
        Pobr.narzedzie = s.narzedzie
        Pobr.pracownik = Prac
        Pobr.save()
        #Pobr.data_pobrania = 

    return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk_prac+"/")
#========================================
#SZABLONY
@login_required(login_url='/narzedziownia/login2/')
def szablony_wybor(request, *args, **kwargs):
    context = {
    }
    url='narzedziownia/szablony_wybor.html'
    dzialy = Pracownik.objects.values('dzial').distinct()
    warianty = Szablon.objects.values('wariant','dzial').distinct()
    print(warianty)
    
    context['dzialy']=dzialy
    context['warianty']=warianty
    return render(request,url,context)
    
@login_required(login_url='/narzedziownia/login2/')
def szablony_lista(request, *args, **kwargs):
    from django.db.models import Q
    context = {
    }
    popraw=False
    dzial=kwargs['dzial']
    wariant=kwargs['wariant']
    wariant_=wariant+" "
    url='narzedziownia/szablony_lista.html'

    Szabl = Szablon.objects.filter(Q(dzial=dzial,wariant=wariant)|Q(dzial=dzial,wariant=wariant_))
    Warianty=Szablon.objects.filter(dzial=dzial).values('wariant').distinct()
    dzialy = Pracownik.objects.values('dzial').distinct()
    context['dzial']=dzial
    context['wariant']=wariant
    context['Szablon']=Szabl
    context['Warianty']=Warianty
    context['dzialy']=dzialy
    if request.method == 'POST':
        if request.POST['narzedzie']!="--------" and request.POST['narzedzie']!="":
            Szabl=Szablon()
            Szabl.dzial=dzial
            Szabl.narzedzie=Narzedzie.objects.get(pk=request.POST['narzedzie'])
            Szabl.ilosc=request.POST['ilosc']
            Szabl.wariant = wariant
            Szabl.save()
        else:
            popraw=True
    else:
        None
    form = SzablonForm()
        
    context['form']=form
    context['popraw']=popraw




    return render(request,url,context)
def szablony_usun(request, *args, **kwargs):
    pk=kwargs['pk']
    dzial=kwargs['dzial']
    wariant=kwargs['wariant']
    print(dzial)
    Szablon.objects.get(pk=pk).delete()
    return HttpResponseRedirect("/narzedziownia/szablony_lista/"+dzial+"/"+wariant+"/")
    
#========================================
#Employer Login
def awaiting(request, *args, **kwargs):
    context={
    }
    p=Potwierdz.objects.first()
    if p.prac!=0:
        pracownik=p.prac
        p.prac=0
        p.save()
        return HttpResponseRedirect("/narzedziownia/pracownicy_view/"+str(pracownik)+"/")
    else:
        url='narzedziownia/awaiting.html'
    return render(request,url,context)

def potwierdz(request, *args, **kwargs):
    pk = kwargs['pk']
    print(pk)
    p=Potwierdz.objects.first()
    p.prac=pk
    p.save()
    return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk+"/")
    
def potwierdz_barcode(request, *args, **kwargs):
    pk = kwargs['pk']
    print(pk)
    p=Potwierdz.objects.first()
    p.prac=pk
    p.save()
    return HttpResponseRedirect("/narzedziownia/pracownicy_edit_barcode/"+pk+"/")
def confirmed(request, *args, **kwargs):
    context={
    }
    
    url='narzedziownia/confirmed.html'

    return render(request,url,context)

def notconfirmed(request, *args, **kwargs):
    context={
    }
    
    url='narzedziownia/notconfirmed.html'

    return render(request,url,context)

    
@login_required(login_url='/narzedziownia/login2/')
def login1(request, *args, **kwargs):
    context={
    }
    kwargs={
    }
    pk=""
    url='narzedziownia/login.html'
    if request.method == "GET":
        form = loginForm(request.GET)
        if form.is_valid():
            print(form.cleaned_data['field'])

            nr_karty=form.cleaned_data['field']
            #nr_karty="172485"
            pk=Pracownik.objects.get(nr_karty=nr_karty).pk
                
            kwargs['pk']=pk
            #pracownicy_edit(request,*args,**kwargs)
            return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+str(pk)+"/")
    context['pk']=pk
    context['form']=form
    return render(request,url,context)
#def signature(request):
#    assert isinstance(request, HttpRequest)
#    url = 'narzedziownia/sign_test.html'
#    context = {
#        'title':'About',
#        'message':'Your application description page.',
#        #'year':datetime.now().year,
#    }
#    return render(request,url,context)
def login2(request, *args, **kwargs):
    context = {
    }
    url='narzedziownia/login2.html'
    if request.method == 'POST':
        form = LoginForm2(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            user = data['user']
            password = data['password']
            user_check = authenticate(username=user, password=password)
            if user_check:
                if user_check.is_authenticated:
                    login(request, user_check)
                    #return render(request,url,context)
                    return HttpResponseRedirect('/narzedziownia/pracownicy_lista/')
                    #HttpResponseRedirect('narzedziownia/pracownicy_lista/')
            else:
                form = LoginForm2()
                context['form']=form
            
                return render(request,url,context)        
    else:
        None
        form = LoginForm2()
    context['form']=form
    return render(request,url,context)


def logout2(request):
    logout(request)
    return HttpResponseRedirect('/narzedziownia/login2')

#                return render(request, 'registration/login.html', {'form': form, 'error': 'Logowanie nie poprawne'})
#    return render(request, 'registration/login.html', {'form': form, 'error': 'Logowanie nie poprawne'})



#    if request.method == 'POST':
#        r=request.user
#        if request.user.is_authenticated():
#    #        print("ffff")
#            return HttpResponseRedirect('/narzedziownia/pracownicy_lista/')
#        else:
#            print("vvvvv")
#        context['r']=r

def sign_test(request, *args, **kwargs):
    url= 'narzedziownia/sign_test.html'
    context={
    }

    form = SignatureForm(request.POST or None)

    if form.is_valid():

        signature = form.cleaned_data.get('signature')
        print(signature)

        if signature:

            #as an image

            signature_picture = draw_signature(signature)

            #or as a file

            signature_file_path = draw_signature(signature, as_file=True)
    context['form']=form
    return render(request,url, context)
def canvas(request):
    url='narzedziownia/canvas.html'
    context = {
    }
    return render(request,url,context)
def przelicz(request,*args, **kwargs):
    from django.db.models import Sum
    pk_prac=kwargs['pk_prac']
    edit=kwargs['edit']
    print(edit)

    for i in Pobranie.objects.filter(pracownik=pk_prac).distinct('narzedzie','data_pobrania'):    #usunac data
        ilosc = Pobranie.objects.filter(pracownik=pk_prac,narzedzie=i.narzedzie,data_pobrania=i.data_pobrania).aggregate(Sum('ilosc')) #usunac data
        p = Pobranie.objects.filter(pracownik=pk_prac,narzedzie=i.narzedzie,data_pobrania=i.data_pobrania) #usunac data
        for k in p:
            k.ilosc=ilosc['ilosc__sum']
            k.save()
        

    for i in Pobranie.objects.filter(pracownik=pk_prac):
        if Pobranie.objects.filter(pracownik=pk_prac,narzedzie=i.narzedzie,data_pobrania=i.data_pobrania).count()>1: #usunac data
            i.delete()
    print(request)
    if edit==0:
        return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk_prac+"/") ###DOROBIC przekierowanie na barcode
    else:
    
        return HttpResponseRedirect("/narzedziownia/pracownicy_edit_barcode/"+pk_prac+"/") ###DOROBIC przekierowanie na barcode


def pdf_odziez(request,*args, **kwargs):
    # -*- coding: utf-8 -*-
    from reportlab.graphics.barcode import qr, eanbc
    from reportlab.graphics.shapes import Drawing
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    from reportlab.graphics import renderPDF
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Praca z PDF
    prac=kwargs['prac']
    #from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger

    pdf_file = 'pdf_odziez.pdf'
     
    can = canvas.Canvas(pdf_file)
    pdfmetrics.registerFont(TTFont('AlegreyaSC', 'apps/narzedziownia/static/narzedziownia/fonts/Alegreya_SC/AlegreyaSC-Bold.ttf'))

    row=1
    for i in PobranieOdziez.objects.filter(pracownik__pk=prac):
        print(can.getAvailableFonts())
        can.setFont("AlegreyaSC",16)
        can.drawString(2*cm , 29*cm - row*cm, i.odziez.nazwa)
        can.drawString(10*cm , 29*cm - row*cm, str(i.ilosc))

        can.drawString(15*cm , 29*cm - row*cm, str(i.data_pobrania))
        row+=1
    can.showPage()
    can.save()

    return HttpResponse("fff")
def zwolniony_alert(request, *args, **kwargs):
    url = 'narzedziownia/zwolniony_alert.html'
    pk = kwargs['pk']
    prac = Pracownik.objects.get(pk=pk)

    context= {
    }
    context['prac']=prac

    
    return render(request,url,context)
@login_required(login_url='/narzedziownia/login2/')
def export(request, *args, **kwargs):
    from openpyxl.styles import colors, Font, Color, Side, Border, PatternFill, GradientFill, Alignment
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    row = 1
    ws['A1']="PRACOWNIK"
    ws['B1']="DZIAL"
    ws['C1']="ODZIEZ(OBUWIE)"
    ws['D1']="DATA POBRANIA"


    for i in PobranieOdziez.objects.filter(odziez__nazwa__contains="OBUWIE").order_by('pracownik__dzial'):
        row += 1
        cellA="A"+str(row)
        ws[cellA]=str(i.pracownik.nazwisko_imie)
        cellB="B"+str(row)
        ws[cellB]=str(i.pracownik.dzial)
        cellC="C"+str(row)
        ws[cellC]=str(i.odziez.nazwa)
        cellD="D"+str(row)
        ws[cellD]=str(i.data_pobrania)
    wb.save("apps/narzedziownia/static/export_odziez.xlsx")

    context = {
    }
    url = 'narzedziownia/export.html'

    return render(request,url,context)
