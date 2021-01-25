from django.shortcuts import render
from .models import *
from .forms import loginForm,PracownikFormConf
from django.http import HttpResponseRedirect
#import pymysql.cursors
#from django.http import httpresponse

def start(request):
    return None
#NARZEDZIA
def narzedzia_lista(request, *args, **kwargs):
    context = {
    }
    url='narzedziownia/narzedzia_lista.html'
    Narzedzia=Narzedzie.objects.all()
    print(Narzedzia)
    context['Narzedzia']=Narzedzia   
    return render(request,url,context)
def narzedzia_edit(request, *args, **kwargs):
    print(request.POST)
    context = {
    }
    url='narzedziownia/narzedzia_edit.html'
    pk=kwargs['pk']
    Narz=Narzedzie.objects.get(pk=pk)
    if request.method == 'POST':
        form = NarzedzieForm(request.POST,instance=Narz)
        if form.is_valid():
            Narz.nazwa=form.cleaned_data['nazwa']
            Narz.opis=form.cleaned_data['opis']
            Narz.save()
        else:
            None
    else:
        form = NarzedzieForm()
    context['form']=form
    context['Narz']=Narz
       #context['form']=form
    return render(request,url,context)
def narzedzia_usun(request, *args, **kwargs):
    pk=kwargs['pk']
    Narzedzie.objects.get(pk=pk).delete()
    return HttpResponseRedirect("/narzedziownia/narzedzia_lista/")
def narzedzia_new(request, *args, **kwargs):
    print(request.POST)
    context = {
    }
    url='narzedziownia/narzedzia_new.html'
    Narz=Narzedzie()
    if request.method == 'POST':
        form = NarzedzieForm(request.POST)
        if form.is_valid():
            print("ffff")
            print(form.cleaned_data['nazwa'])

            Narz.nazwa=form.cleaned_data['nazwa']
            Narz.opis=form.cleaned_data['opis']
            Narz.save()
            return HttpResponseRedirect("/narzedziownia/narzedzia_lista/")
        else:
            None
    else:
        form = NarzedzieForm()
    context['form']=form
    context['Narz']=Narz
    context['var']=1
       #context['form']=form
    return render(request,url,context)
#EMPLOYEE/PRACOWNICY

def pracownicy_lista(request, *args, **kwargs):
    import MySQLdb
    #=== DODAJ ZATRUDNIENIA
    try:
        conn=MySQLdb.connect(host="192.168.41.15",user="kadry", passwd="start", db="KADRY",port=3306,charset='utf8')
        cur = conn.cursor()
        sql = "SELECT * FROM Nowi_pracownicy" 
        cur.execute(sql)
        rows=cur.fetchall()
    except mdb.Error:
        print("ERROR")
        #print "Error %d: %s" % (e.args[0],e.args[1])
        sys.exit(1)
    finally:
        if conn:
            conn.close()
    for i in rows:
        Prac, created=Pracownik.objects.get_or_create(nazwisko_imie="{} {}".format(i[2],i[1]),dzial="{}".format(i[3]).upper(),stanowisko="{}".format(i[4]))

        #Prac=Pracownik()
        #Prac.nazwisko_imie="{} {}".format(i[2],i[1])
        #Prac.dzial="{}".format(i[3])
        #Prac.stanowisko="{}".format(i[4])
        #Prac.save()
        
        #print(i[0])
        #print(i[1])
        #print(i[2])
        #print(i[3])
        #print(i[4])
        
        

    context = {
    }
    url='narzedziownia/pracownicy_lista.html'
    Pracownicy=Pracownik.objects.all()
    context['Pracownicy']=Pracownicy
    return render(request,url,context)

def pracownicy_edit_(request, *args, **kwargs):
    from collections import OrderedDict
    context = {
    }
    url='narzedziownia/pracownicy_edit.html'
    pk=kwargs['pk']
    #print(kwargs)
    Numery_narz=range(10)
    Prac=Pracownik.objects.get(pk=pk)
    Narz_temp_str=Prac.narzedzia
    if Narz_temp_str!=None:
        Narz_temp_list=Narz_temp_str.split(";")
        Narz_temp=OrderedDict() 
        print(Narz_temp_list)

        for i in Narz_temp_list:
            if i.split(":")[0]!="":  #wuwalamy ostatniego pustaka
                try:
                    Narz_temp[i.split(":")[0]]=i.split(":")[1]
                except:
                    break;
    else:
        Narz_temp=[]
    

    Lista_narz=Narzedzie.objects.all()
    print(request.POST)
    narz_list=""
    if request.method == 'POST':
        form = PracownikForm(request.POST,instance=Prac)
        if form.is_valid():
            Prac.komentarz=form.cleaned_data['komentarz']
            for i in Numery_narz:
                narz="narz_{}".format(i)
                #is_narz="is_narz_{}".format(i)
                il_narz="il_narz_{}".format(i)
                narz_list+=request.POST[narz]
                narz_list+=":"
                narz_list+=request.POST[il_narz]
                narz_list+=";"
            Prac.narzedzia=narz_list
            Prac.save()
            return HttpResponseRedirect("/narzedziownia/pracownicy_lista/")
        else:
            None
    else:
        form =PracownikForm()
    context['form']=form
    context['Prac']=Prac
    context['Numery_narz']=Numery_narz
    


    context['Lista_narz']=Lista_narz
    context['Narz_temp']=Narz_temp
       #context['form']=forma
    print(Lista_narz,Narz_temp)
    return render(request,url,context)

def pracownicy_usun_pobranie(request, *args, **kwargs):
    #context = {
    #}
    #url='narzedziownia/pracownicy_usun_pobranie.html'
    pk_pobr=kwargs['pk_pobr']
    pk_prac=kwargs['pk_prac']
    Pobr=Pobranie.objects.filter(pk=pk_pobr).delete()
    return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk_prac+"/")

def pracownicy_edit(request, *args, **kwargs):
    context = {
    }
    url='narzedziownia/pracownicy_edit.html'
    pk=kwargs['pk']
    #print(kwargs)
    Prac=Pracownik.objects.get(pk=pk)
    Narz_temp_str=Prac.narzedzia

    Lista_narz=Narzedzie.objects.all()
    narz_list=""
    Pobr=Pobranie()
    if request.method == 'POST':

        pk_narz=request.POST['narzedzie']
        Pobr.pracownik=Prac
        Pobr.narzedzie=Narzedzie.objects.get(pk=pk_narz)
        Pobr.ilosc=request.POST['ilosc']
        if request.POST['data_pobrania']!="":
            Pobr.data_pobrania=request.POST['data_pobrania']
        if request.POST['data_oddania']!="":
            print("sssss")
            Pobr.data_oddania=request.POST['data_oddania']
        Pobr.save()
        return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk+"/")
        #else:
        #    None
    else:
        form = PobranieForm()
        #form =PracownikForm()a
    Pobr_filter=Pobranie.objects.filter(pracownik=pk)
    context['Pobr_filter']=Pobr_filter
    context['form']=form
    context['Prac']=Prac
    context['Lista_narz']=Lista_narz
       #context['form']=forma
    return render(request,url,context)

def pracownicy_view(request, *args, **kwargs):
    context = {
    }
    url='narzedziownia/pracownicy_view.html'
    pk=kwargs['pk']
    #print(kwargs)
    Prac=Pracownik.objects.get(pk=pk)
    Narz_temp_str=Prac.narzedzia


    Lista_narz=Narzedzie.objects.all()
    narz_list=""
    Pobr=Pobranie()
    if request.method == 'GET':
        form = PracownikFormConf(request.GET)
        if form.is_valid():
            print(form.cleaned_data['potwierdzenie'])
            Prac.potwierdzenie=form.cleaned_data['potwierdzenie']
            Prac.save()
            return HttpResponseRedirect("/narzedziownia/confirmed/")
        else:
            None
    else:
        form =PracownikFormConf()
    context['form']=form
    context['Prac']=Prac
    Pobr_filter=Pobranie.objects.filter(pracownik=pk)
    Pobr_filter_kol1=Pobranie.objects.filter(pracownik=pk)[:9]
    Pobr_filter_kol2=Pobranie.objects.filter(pracownik=pk)[10:19]
    Pobr_filter_kol3=Pobranie.objects.filter(pracownik=pk)[20:]
        
    context['Pobr_filter']=Pobr_filter
    context['Pobr_filter_kol1']=Pobr_filter_kol1
    context['Pobr_filter_kol2']=Pobr_filter_kol2
    context['Pobr_filter_kol3']=Pobr_filter_kol3
    return render(request,url,context)

def pracownicy_view_(request, *args, **kwargs):
    from collections import OrderedDict
    context = {
    }
    url='narzedziownia/pracownicy_view.html'
    pk=kwargs['pk']
    #print(kwargs)
    Numery_narz=range(10)
    Prac=Pracownik.objects.get(pk=pk)
    Narz_temp_str=Prac.narzedzia
    if Narz_temp_str!=None:
        Narz_temp_list=Narz_temp_str.split(";")
        Narz_temp=OrderedDict() 
        print(Narz_temp_list)

        for i in Narz_temp_list:
            if i.split(":")[0]!="":  #wuwalamy ostatniego pustaka
                try:
                    Narz_temp[i.split(":")[0]]=i.split(":")[1]
                except:
                    break;
    else:
        Narz_temp=[]
    

    Lista_narz=Narzedzie.objects.all()
    print(request.POST)
    narz_list=""
    if request.method == 'GET':
        form = PracownikFormConf(request.GET)
        if form.is_valid():
            print("BLAAA")
            print(form.cleaned_data['potwierdzenie'])
            Prac.potwierdzenie=form.cleaned_data['potwierdzenie']

            #Prac.potwierdzenie=request.POST('potwierdzenie')
            
            #Prac.komentarz=form.cleaned_data['komentarz']
            #for i in Numery_narz:
            #    narz="narz_{}".format(i)
            #    #is_narz="is_narz_{}".format(i)
            #    il_narz="il_narz_{}".format(i)
            #    narz_list+=request.POST[narz]
            #    narz_list+=":"
            #    narz_list+=request.POST[il_narz]
            #    narz_list+=";"
            #Prac.narzedzia=narz_list
            

            Prac.save()
            return HttpResponseRedirect("/narzedziownia/login/")
        else:
            None
    else:
        form =PracownikFormConf()
    context['form']=form
    context['Prac']=Prac
    context['Numery_narz']=Numery_narz
    


    context['Lista_narz']=Lista_narz
    context['Narz_temp']=Narz_temp
       #context['form']=form
    return render(request,url,context)

def import_pracownikow(request, *args, **kwargs):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    import os
    url='narzedziownia/import_pracownikow.html'
    context={
    }

    path_src="Stan.XLSX"
    wb_src = load_workbook(path_src) 
    sheet_src = wb_src.active
    rows=[]
    
    for i in range(1500)[2:]:
        row = []
        cell_src_1 = sheet_src.cell(row = i, column = 1)
        cell_src_2 = sheet_src.cell(row = i, column = 2)
        cell_src_3 = sheet_src.cell(row = i, column = 3)
        cell_src_4 = sheet_src.cell(row = i, column = 4)
        cell_src_5 = sheet_src.cell(row = i, column = 5) #DATA
        cell_src_6 = sheet_src.cell(row = i, column = 6) #SPOT
        cell_src_7 = sheet_src.cell(row = i, column = 7) #Rodzaj wynagrodzenia
        cell_src_8 = sheet_src.cell(row = i, column = 8) #TypUmowy
        cell_src_9 = sheet_src.cell(row = i, column = 9) #Karta
        try:
            p, created=Pracownik.objects.get_or_create(nr_pracownika=cell_src_1.value)
            p.nr_pracownika=cell_src_1.value
            p.nazwisko_imie=cell_src_2.value
            p.dzial=cell_src_3.value
            p.stanowisko=cell_src_4.value
            p.zatrudnienie=cell_src_5.value
            p.spot=cell_src_6.value
            p.podgrupa_prac=cell_src_7.value
            p.typ_umowy=cell_src_8.value
            p.nr_karty=int(cell_src_9.value)
            p.save()
            rows.append(p)
        except:
            None
    context['rows']=rows
    return render(request,url,context)
def pracownicy_szablon(request, *args, **kwargs):
    pk_prac = kwargs['pk_prac']
    Prac=Pracownik.objects.get(pk=pk_prac)
    Dzial=Prac.dzial
    for s in Szablon.objects.filter(dzial=Dzial):
        Pobr =  Pobranie()
        Pobr.ilosc=s.ilosc
        Pobr.narzedzie = s.narzedzie
        Pobr.pracownik = Prac
        Pobr.save()
        #Pobr.data_pobrania = 

    return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk_prac+"/")
#========================================
#SZABLONY
def szablony_wybor(request, *args, **kwargs):
    context = {
    }
    url='narzedziownia/szablony_wybor.html'
    dzialy = Pracownik.objects.values('dzial').distinct()
    warianty = Szablon.objects.values('wariant','dzial').distinct()
    print(warianty)
    
    context['dzialy']=dzialy
    context['warianty']=warianty
    return render(request,url,context)
    
def szablony_lista(request, *args, **kwargs):
    context = {
    }
    dzial=kwargs['dzial']
    wariant=kwargs['wariant']
    url='narzedziownia/szablony_lista.html'
    Szabl = Szablon.objects.filter(dzial=dzial,wariant=wariant)
    dzialy = Pracownik.objects.values('dzial').distinct()
    context['dzial']=dzial
    context['Szablon']=Szabl
    context['dzialy']=dzialy
    if request.method == 'POST':
        Szabl=Szablon()
        Szabl.dzial=dzial
        Szabl.narzedzie=Narzedzie.objects.get(pk=request.POST['narzedzie'])
        Szabl.ilosc=request.POST['ilosc']

        Szabl.save()
    else:
        form = SzablonForm()
        
        context['form']=form




    return render(request,url,context)
def szablony_usun(request, *args, **kwargs):
    pk=kwargs['pk']
    dzial=kwargs['dzial']
    print(dzial)
    Szablon.objects.get(pk=pk).delete()
    return HttpResponseRedirect("/narzedziownia/szablony_lista/"+dzial+"/")
    
#========================================
#Employer Login
def awaiting(request, *args, **kwargs):
    context={
    }
    p=Potwierdz.objects.first()
    if p.prac!=0:
        pracownik=p.prac
        p.prac=0
        p.save()
        return HttpResponseRedirect("/narzedziownia/pracownicy_view/"+str(pracownik)+"/")
    else:
        url='narzedziownia/awaiting.html'
    return render(request,url,context)

def potwierdz(request, *args, **kwargs):
    pk = kwargs['pk']
    print(pk)
    p=Potwierdz.objects.first()
    p.prac=pk
    p.save()
    return HttpResponseRedirect("/narzedziownia/pracownicy_edit/"+pk+"/")
    
def confirmed(request, *args, **kwargs):
    context={
    }
    
    url='narzedziownia/confirmed.html'

    return render(request,url,context)


    
def login(request, *args, **kwargs):
    context={
    }
    kwargs={
    }
    pk=""
    url='narzedziownia/login.html'
    if request.method == "GET":
        form = loginForm(request.GET)
        if form.is_valid():
            print(form.cleaned_data['field'])

            nr_karty=form.cleaned_data['field']
            #nr_karty="172485"
            pk=Pracownik.objects.get(nr_karty=nr_karty).pk
                
            kwargs['pk']=pk
            #pracownicy_edit(request,*args,**kwargs)
            return HttpResponseRedirect("/narzedziownia/pracownicy_view/"+str(pk)+"/")
    context['pk']=pk
    context['form']=form
    return render(request,url,context)
